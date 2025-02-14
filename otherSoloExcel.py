import xlrd
import os
from openpyxl import Workbook,load_workbook
import pandas as pd
from tkinter import messagebox

# 主函数：遍历用户选择的文件夹中的所有Excel文件并处理
# xls_file为用户选择的文件夹路径，first_search_text为第一个搜索文本，sheet_name为工作表名称
def search_and_output_data(xls_file, first_search_text, output_file_path): 
    # 打开Excel文件
    workbook = xlrd.open_workbook(xls_file)

    # 用户手动输入要处理的工作表
    sheet = workbook.sheet_by_name('客户交易结算日报')  # 或者使用 sheet_by_name('Sheet1')

    # 打开通过create_excel()创建的汇总.xlsx文件
    wb = load_workbook(output_file_path + '汇总.xlsx')
    ws=wb.active

    # 遍历所有行以找到第一个搜索文本
    for row_idx in range(sheet.nrows):
        if sheet.cell_value(row_idx, 0) == first_search_text:  # 假设在A列搜索
            # 确定第一个搜索文本的行号，并从下一行开始搜索第二个文本
            start_row_for_second_search = row_idx + 1
            break  # 找到第一个文本后停止搜索
    else:
        # 如果没有找到第一个搜索文本，则输出错误消息并退出
        print(f"Search text '{first_search_text}' not found in the workbook.")
        return

    # 从指定的行开始搜索第二个搜索文本
    for row_idx in range(start_row_for_second_search + 1, sheet.nrows):
        cell_value = sheet.cell_value(row_idx, 0)
        # 判断单元格的值是否等于search_text
        if cell_value == "合计":
            print("遍历完成")
            break
        row_values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(5)]
        print(row_values) # 输出数据
        for output_row_idx in range(2,1000): # 2到1000行进行数据写入，第一行为表头
            if all(ws.cell(row=output_row_idx, column=col_idx + 1).value is None for col_idx in range(5)):
                # 如果该行没有数据，则写入数据
                for col_idx, value in enumerate(row_values):
                    ws.cell(row=output_row_idx, column=col_idx + 1, value=value)
                break
            else:
                continue
    wb.save(os.path.join(output_file_path, '汇总.xlsx')) # 保存文件

## 定义遍历文件夹的函数
def traverse_folder_and_test(folder_path, first_search_text, output_file_path):
    # 使用os.listdir()列出文件夹中的所有文件和文件夹
    all_items = os.listdir(folder_path)
    
    # 遍历文件夹中的每一项
    for item in all_items:
        # 获取完整路径
        full_path = os.path.join(folder_path, item)
        
        # 检查是否为文件
        if os.path.isfile(full_path):
            # 如果是文件，则调用test方法
            search_and_output_data(full_path, first_search_text,output_file_path)

# 读取保存的output.xlsx文件并汇总求和
def sum_xlsx_columns(output_file_path):
    # 读取Excel文件
    df = pd.read_excel(output_file_path + '汇总.xlsx')
    # 使用groupby根据第一列对数据进行分组，然后对第2、3、4、5列进行求和
    sums = df.groupby(df.columns[0])[df.columns[1:5]].sum()
    # 重置索引，使分组后的数据重新成为一个DataFrame
    sums.reset_index(inplace=True)
    
    print(sums)
    sums.to_excel(output_file_path + '汇总.xlsx', index=False)  # 将结果保存到Excel文件中
    messagebox.showinfo("提示", "汇总.xlsx生成成功")

# 创建汇总.xlsx文件
def create_or_open_excel(folder_path):
    # 如果文件不存在，则创建一个新的工作簿
    wb = Workbook()
    # 选择活动的工作表
    ws = wb.active
    # 在第一行的五列中分别插入指定的标题
    ws.append(["合约", "买持仓", "买均价", "卖持仓", "卖均价"])
    # 保存新的Excel文件
    wb.save(os.path.join(folder_path, '汇总.xlsx'))

    # 如果文件已存在，我们仍然需要检查第一行是否包含正确的标题
    ws = wb.active
    if ws['A1'].value != "合约":
        # 如果第一列的标题不正确，则更新它和其它列的标题
        ws['A1'] = "合约"
        ws['B1'] = "买持仓"
        ws['C1'] = "买均价"
        ws['D1'] = "卖持仓"
        ws['E1'] = "卖均价"
        # 保存更新后的Excel文件
        wb.save(os.path.join(folder_path, '汇总.xlsx'))

# # 定义变量
# first_search_text = '期货持仓汇总'  # 替换为你要搜索的第一个文字
# xls_file = 'xlsfile' # 文件夹
# #xls_file = '2024-11-14  东证  永盛  00000000000_2024-11-14.xls'  # 替换为你的Excel文件路径

# # 调用主函数
# traverse_folder_and_test(xls_file,first_search_text)


